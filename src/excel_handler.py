"""Excel (xlsx/xls) 导入/导出处理。"""

from __future__ import annotations

import re
from pathlib import Path

from src.scanner import expand_ip_range, expand_port_range


def _is_header_row(values: list[str]) -> bool:
    """判断一行是否为标题行。"""
    if not values:
        return False
    first = str(values[0]).strip().lower()
    keywords = {"ip", "地址", "address", "host", "主机", "端口", "port", "描述", "description", "集合", "collection"}
    return first in keywords or any(kw in first for kw in keywords)


# ── 导入 ───────────────────────────────────────────────────


def parse_targets_excel(filepath: str | Path) -> tuple[list[dict], list[str]]:
    """从 Excel 文件导入目标。

    列顺序: IP地址, 端口, 描述, 集合名称 (第一行为标题则跳过)

    Returns:
        (targets, errors) — targets 为 [{"ip","port","description","collection_name"}, ...]
    """
    filepath = Path(filepath)
    ext = filepath.suffix.lower()
    targets: list[dict] = []
    errors: list[str] = []

    rows_data: list[list] = []

    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(filepath))
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            rows_data.append([ws.cell_value(r, c) for c in range(ws.ncols)])
    else:
        from openpyxl import load_workbook
        wb = load_workbook(str(filepath), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows_data.append(list(row))
        wb.close()

    for i, row in enumerate(rows_data, start=1):
        # 跳过完全空行
        values = [str(c).strip() if c is not None else "" for c in row]
        if all(v == "" for v in values):
            continue

        # 首行可能是标题
        if i == 1 and _is_header_row(values):
            continue

        if len(values) < 2:
            errors.append(f"第 {i} 行: 列数不足（需要 IP 和端口）")
            continue

        ip_raw = str(values[0]).strip() if values[0] is not None else ""
        port_raw = str(values[1]).strip() if values[1] is not None else ""
        desc = str(values[2]).strip() if len(values) > 2 and values[2] is not None else ""
        collection = str(values[3]).strip() if len(values) > 3 and values[3] is not None else ""

        # 展开 IP（换行 + CIDR/范围）
        raw_ips = [ip.strip() for ip in ip_raw.split("\n") if ip.strip()]
        if not raw_ips:
            errors.append(f"第 {i} 行: IP 地址为空")
            continue

        ips = []
        for raw_ip in raw_ips:
            try:
                ips.extend(expand_ip_range(raw_ip))
            except ValueError as e:
                errors.append(f"第 {i} 行: {e}")
                continue

        # 展开端口（换行 + 范围）
        raw_ports = [p.strip() for p in port_raw.split("\n") if p.strip()]
        if not raw_ports:
            # 兼容旧格式：单行逗号/范围格式
            raw_ports = [port_raw]
        ports = []
        for rp in raw_ports:
            try:
                ports.extend(expand_port_range(rp))
            except ValueError as e:
                errors.append(f"第 {i} 行: {e}")
                continue

        if not ports:
            errors.append(f"第 {i} 行: 端口无效")
            continue

        # 笛卡尔积
        for ip in ips:
            for port in ports:
                targets.append({"ip": ip, "port": port, "description": desc, "collection_name": collection})

    return targets, errors


# ── 导出 ───────────────────────────────────────────────────


def _format_row(values: list) -> list:
    """确保单元格值都是基础类型（非 numpy 等）。"""
    return [float(v) if isinstance(v, (int, float)) else str(v) for v in values]


def _border_all(ws, max_row: int, max_col: int):
    """给指定范围的所有单元格添加细线边框。"""
    from openpyxl.styles import Border, Side
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = border


def export_targets_to_excel(filepath: str | Path, targets: list[dict]) -> tuple[bool, str]:
    """导出目标列表到 Excel (.xlsx)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "目标列表"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        headers = ["IP地址", "端口", "描述", "集合", "创建时间"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r, t in enumerate(targets, 2):
            ws.cell(row=r, column=1, value=t.get("ip", ""))
            ws.cell(row=r, column=2, value=int(t.get("port", 0)))
            ws.cell(row=r, column=3, value=t.get("description", ""))
            ws.cell(row=r, column=4, value=t.get("collection_name", ""))
            ws.cell(row=r, column=5, value=t.get("created_at", ""))

        _border_all(ws, len(targets) + 1, len(headers))

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 20

        wb.save(str(filepath))
        return True, ""
    except Exception as e:
        return False, str(e)


# XML 非法字符（除 TAB/LF/CR 外的 C0 控制字符）与孤立代理对。
# 前者会让 Office 严格校验时拒绝打开文件，后者会在 UTF-8 编码时抛异常。
_ILLEGAL_CHARS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")
_SURROGATES_RE = re.compile(r"[\ud800-\udfff]")


def _sanitize_text(value):
    """将字符串中的 XML 非法字符替换为可读的 \\xNN / \\uNNNN 转义形式。

    保留 TAB/LF/CR（\t、\n、\r，XML 合法）。非字符串值原样返回。
    响应/错误报文可能含二进制协议数据解码出的控制字符，直接写入
    xlsx 会导致导出失败（openpyxl IllegalCharacterError）或生成
    Office 打不开的文件。
    """
    if not isinstance(value, str):
        return value
    value = _ILLEGAL_CHARS_RE.sub(
        lambda m: f"\\x{ord(m.group(0)):02x}", value)
    value = _SURROGATES_RE.sub(
        lambda m: f"\\u{ord(m.group(0)):04x}", value)
    return value


def export_rows_to_excel(filepath: str | Path, headers: list[str],
                         rows: list[list]) -> tuple[bool, str]:
    """通用导出：将表头 + 数据行导出为 .xlsx（带表头样式与边框）。

    rows 中的单元格值应为 str/int/float/bool 等基础类型。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "导出数据"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=_sanitize_text(val))

        _border_all(ws, len(rows) + 1, len(headers))
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 20

        wb.save(str(filepath))
        return True, ""
    except Exception as e:
        return False, str(e)


def export_results_to_excel(filepath: str | Path, results: list[dict]) -> tuple[bool, str]:
    """导出测试结果到 Excel (.xlsx)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "测试结果"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        headers = ["IP地址", "端口", "描述", "集合", "状态", "延迟(ms)", "错误信息", "检测时间"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r, t in enumerate(results, 2):
            success = t.get("success", False)
            row_fill = green_fill if success else red_fill

            for c, val in enumerate([
                t.get("ip", ""),
                int(t.get("port", 0)),
                t.get("description", ""),
                t.get("collection_name", ""),
                "连通" if success else "未连通",
                round(float(t.get("latency_ms", 0)), 1) if success else "",
                t.get("error_msg", ""),
                t.get("tested_at", ""),
            ], 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill = row_fill

        _border_all(ws, len(results) + 1, len(headers))

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 20

        wb.save(str(filepath))
        return True, ""
    except Exception as e:
        return False, str(e)
